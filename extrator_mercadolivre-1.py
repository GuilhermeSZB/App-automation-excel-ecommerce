"""
Extrator de Dados - Mercado Livre / Mercado Shops
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import os

# ─────────────────────────────────────────────
# COLUNAS DO MERCADO LIVRE → nome na planilha nova
# ─────────────────────────────────────────────
COLUNAS_DESEJADAS = {
    "Título do anúncio":                "Nome do produto",
    "Receita por produtos (BRL)":       "Receita",
    "Tarifa de venda e impostos (BRL)": "Tarifa 12% ou 17%",
    "Receita por envio (BRL)":          "Receita por envio",
    "Tarifas de envio (BRL)":           "Tarifa de envio",
    "Descontos e bônus":                "Desconto bônus",
    "Total (BRL)":                      "Total sobre",
    "Estado.1":                         "Estado",
}

COLUNA_ESTADO       = "Estado"
COLUNA_TAXA_ESCOLHA = "Taxa Produto"      # dropdown: 7%, 8,8%, Sem taxa
COLUNA_TAXA_VALOR   = "Valor Taxa"        # calculado automaticamente
COLUNA_ICMS         = "Taxa 6,06%"        # sobre Total sobre
COLUNA_GUBEE        = "Taxa Gubee"
PCT_ICMS            = 0.0606
VALOR_GUBEE         = 1.0

OPCOES_TAXA = ["Sem taxa", "7%", "8,8%"]
MAPA_PCT    = {"Sem taxa": 0.0, "7%": 0.07, "8,8%": 0.088}


# ─────────────────────────────────────────────
# HELPERS DE ESTILO (xlsx)
# ─────────────────────────────────────────────

def cor(hex6):
    return Color(rgb="FF" + hex6.upper())

def fill(hex6):
    return PatternFill(fill_type="solid", fgColor=cor(hex6))

def borda_cel():
    c = "AAAAAA"
    return Border(
        left=Side(style="thin", color=c), right=Side(style="thin", color=c),
        top=Side(style="thin", color=c),  bottom=Side(style="thin", color=c),
    )


# ─────────────────────────────────────────────
# PROCESSAMENTO
# ─────────────────────────────────────────────

def preparar_dados(caminho_entrada):
    def encontrar_cab(caminho):
        df_raw = pd.read_excel(caminho, header=None, nrows=20)
        for i, row in df_raw.iterrows():
            vals = [str(v) for v in row.values]
            if any("N.º de venda" in v or "Receita por produtos" in v for v in vals):
                return i
        return 4

    linha_cab = encontrar_cab(caminho_entrada)
    df = pd.read_excel(caminho_entrada, sheet_name=0, header=linha_cab, dtype=str)
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    encontradas = {orig: novo for orig, novo in COLUNAS_DESEJADAS.items() if orig in df.columns}
    faltando    = [orig for orig in COLUNAS_DESEJADAS if orig not in df.columns]

    if not encontradas:
        raise ValueError("Nenhuma coluna conhecida foi encontrada.")

    df_s = df[list(encontradas.keys())].copy()
    df_s.rename(columns=encontradas, inplace=True)
    df_s = df_s.dropna(how="all").reset_index(drop=True)

    # Taxa 6,06% sobre Total sobre
    if "Total sobre" in df_s.columns:
        total = pd.to_numeric(df_s["Total sobre"].str.replace(",", ".", regex=False), errors="coerce").fillna(0)
        df_s[COLUNA_ICMS] = (total * PCT_ICMS).round(2)
    else:
        df_s[COLUNA_ICMS] = 0.0

    # Taxa por produto — inicia como "Sem taxa"
    df_s[COLUNA_TAXA_ESCOLHA] = "Sem taxa"
    df_s[COLUNA_TAXA_VALOR]   = 0.0

    # Gubee fixo
    df_s[COLUNA_GUBEE] = VALOR_GUBEE

    return df_s, faltando


def calcular_valor_taxa(df, row_idx):
    """Recalcula o Valor Taxa de uma linha baseado na escolha e no Total sobre."""
    escolha = df.at[row_idx, COLUNA_TAXA_ESCOLHA]
    pct = MAPA_PCT.get(escolha, 0.0)
    if pct == 0.0:
        return 0.0
    try:
        total = float(str(df.at[row_idx, "Total sobre"]).replace(",", "."))
        return round(total * pct, 2)
    except:
        return 0.0


def salvar_xlsx(df_saida, caminho_saida):
    wb = Workbook()
    wb.loaded_theme = None
    ws = wb.active
    ws.title = "Vendas Extraídas"

    colunas = list(df_saida.columns)
    ws.row_dimensions[1].height = 30

    # Cores dos cabeçalhos
    COR_CAB = {
        COLUNA_ICMS:         "B7410E",   # laranja escuro
        COLUNA_GUBEE:        "1E8449",   # verde
        COLUNA_ESTADO:       "6C3483",   # roxo
        COLUNA_TAXA_ESCOLHA: "1A5276",   # azul médio
        COLUNA_TAXA_VALOR:   "154360",   # azul escuro
    }

    for c_idx, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=c_idx, value=nome)
        cel.fill      = fill(COR_CAB.get(nome, "1B4F72"))
        cel.font      = Font(name="Arial", size=11, bold=True, color=cor("FFFFFF"))
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border    = borda_cel()

    # Cores de fundo das células especiais
    COR_CEL = {
        COLUNA_ICMS:         "FDEBD0",   # laranja claro
        COLUNA_GUBEE:        "D5F5E3",   # verde claro
        COLUNA_ESTADO:       "E8DAEF",   # roxo claro
        COLUNA_TAXA_ESCOLHA: "D6EAF8",   # azul claro fixo
        COLUNA_TAXA_VALOR:   "D1F2EB",   # verde-água claro
    }

    for r_idx, (_, row) in enumerate(df_saida.iterrows(), start=2):
        ws.row_dimensions[r_idx].height = 16
        cor_linha = "D6EAF8" if r_idx % 2 == 0 else "FFFFFF"

        for c_idx, nome_col in enumerate(colunas, start=1):
            v = row[nome_col]

            if nome_col == COLUNA_GUBEE:
                val = VALOR_GUBEE
            elif nome_col in (COLUNA_ICMS, COLUNA_TAXA_VALOR):
                try:    val = float(v)
                except: val = 0.0
            elif nome_col == COLUNA_TAXA_ESCOLHA:
                val = str(v) if str(v) in OPCOES_TAXA else "Sem taxa"
            else:
                try:    val = float(str(v).replace(",", ".")) if str(v) not in ("", "nan", "None") else ""
                except: val = str(v) if str(v) not in ("nan", "None") else ""

            cel = ws.cell(row=r_idx, column=c_idx, value=val)
            cel.font   = Font(name="Arial", size=10)
            cel.border = borda_cel()

            bg = COR_CEL.get(nome_col, cor_linha)
            cel.fill      = fill(bg)
            cel.alignment = Alignment(
                horizontal="left" if nome_col == "Nome do produto" else "center",
                vertical="center"
            )
            if nome_col in (COLUNA_GUBEE, COLUNA_ICMS, COLUNA_TAXA_VALOR):
                cel.number_format = 'R$ #,##0.00'
            elif isinstance(val, float) and nome_col not in (COLUNA_TAXA_ESCOLHA, COLUNA_ESTADO):
                cel.number_format = 'R$ #,##0.00'

    larguras = {
        "Nome do produto":  50,
        COLUNA_GUBEE:       14,
        COLUNA_ICMS:        16,
        COLUNA_ESTADO:      22,
        COLUNA_TAXA_ESCOLHA:16,
        COLUNA_TAXA_VALOR:  16,
    }
    for c_idx, nome_col in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = larguras.get(nome_col, 20)

    ws.freeze_panes = "A2"
    wb.save(caminho_saida)


# ─────────────────────────────────────────────
# PALETA DE CORES DA UI
# ─────────────────────────────────────────────
UI = {
    "bg_escuro":    "#0F2942",
    "bg_claro":     "#F5F7FA",
    "bg_card":      "#FFFFFF",
    "azul":         "#1565C0",
    "azul_hover":   "#0D47A1",
    "verde":        "#2E7D32",
    "verde_hover":  "#1B5E20",
    "borda":        "#DCE3ED",
    "texto":        "#1A2A3A",
    "subtexto":     "#607D8B",
    "tag_bg":       "#E3F2FD",
    "tag_fg":       "#1565C0",
    "tag_gubee":    "#E8F5E9",
    "tag_gubee_fg": "#2E7D32",
}

COR_CAB_UI = {
    COLUNA_ICMS:         "#B7410E",
    COLUNA_GUBEE:        "#1E8449",
    COLUNA_ESTADO:       "#6C3483",
    COLUNA_TAXA_ESCOLHA: "#1A5276",
    COLUNA_TAXA_VALOR:   "#154360",
}
COR_CEL_UI = {
    COLUNA_ICMS:         "#FDEBD0",
    COLUNA_GUBEE:        "#D5F5E3",
    COLUNA_ESTADO:       "#E8DAEF",
    COLUNA_TAXA_ESCOLHA: "#D6EAF8",
    COLUNA_TAXA_VALOR:   "#D1F2EB",
}


def _hover(btn, normal, hover):
    btn.bind("<Enter>", lambda e: btn.config(bg=hover))
    btn.bind("<Leave>", lambda e: btn.config(bg=normal))


# ─────────────────────────────────────────────
# JANELA DE EDIÇÃO DA TABELA
# ─────────────────────────────────────────────

class JanelaTabela(tk.Toplevel):

    def __init__(self, parent, df, caminho_saida, cb_salvo):
        super().__init__(parent)
        self.title("📋  Tabela de Dados – edição")
        self.geometry("1200x580")
        self.configure(bg=UI["bg_claro"])
        self.resizable(True, True)

        self.df            = df.copy()
        self.caminho_saida = caminho_saida
        self.cb_salvo      = cb_salvo
        self._celula_editando = None
        self._cells = {}

        self._construir()

    # ── Layout ──────────────────────────────────

    def _construir(self):
        toolbar = tk.Frame(self, bg=UI["bg_escuro"], height=46)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        tk.Label(toolbar, text="  📋  Tabela de Dados",
            font=("Arial", 12, "bold"), bg=UI["bg_escuro"], fg="white"
        ).pack(side="left", padx=8, pady=10)

        tk.Label(toolbar,
            text="Duplo clique para editar  •  Coluna 'Taxa Produto': escolha a taxa de cada linha  •  'Valor Taxa' é calculado automaticamente",
            font=("Arial", 8), bg=UI["bg_escuro"], fg="#90CAF9"
        ).pack(side="left", padx=10)

        btn_salvar = tk.Button(toolbar, text="  💾  Salvar Planilha  ",
            command=self._salvar,
            bg=UI["verde"], fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=6, pady=4
        )
        btn_salvar.pack(side="right", padx=12, pady=8)
        _hover(btn_salvar, UI["verde"], UI["verde_hover"])

        self.lbl_info = tk.Label(toolbar,
            text=f"  {len(self.df)} linhas  |  {len(self.df.columns)} colunas",
            font=("Arial", 8), bg=UI["bg_escuro"], fg="#90CAF9"
        )
        self.lbl_info.pack(side="right", padx=4)

        container = tk.Frame(self, bg=UI["bg_claro"])
        container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(container, bg=UI["bg_claro"], highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical",   command=self.canvas.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.frame_tabela = tk.Frame(self.canvas, bg=UI["bg_claro"])
        self.canvas_window = self.canvas.create_window((0, 0), window=self.frame_tabela, anchor="nw")
        self.frame_tabela.bind("<Configure>", lambda e: self.canvas.configure(
            scrollregion=self.canvas.bbox("all")))

        self.canvas.bind_all("<MouseWheel>",
            lambda e: self.canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.canvas.bind_all("<Shift-MouseWheel>",
            lambda e: self.canvas.xview_scroll(int(-1*(e.delta/120)), "units"))

        self._renderizar_tabela()

    def _cor_cab(self, nome):
        return COR_CAB_UI.get(nome, "#1B4F72")

    def _cor_cel(self, nome_col, r_idx):
        if nome_col in COR_CEL_UI:
            return COR_CEL_UI[nome_col]
        return "#EAF4FB" if r_idx % 2 == 0 else "#FFFFFF"

    def _largura(self, nome):
        if nome == "Nome do produto":   return 300
        if nome == COLUNA_TAXA_ESCOLHA: return 110
        if nome == COLUNA_TAXA_VALOR:   return 110
        if nome == COLUNA_ESTADO:       return 150
        if nome in (COLUNA_ICMS, COLUNA_GUBEE): return 110
        return 120

    def _renderizar_tabela(self):
        for w in self.frame_tabela.winfo_children():
            w.destroy()
        self._cells = {}

        colunas = list(self.df.columns)

        # Cabeçalho
        tk.Label(self.frame_tabela, text="#", width=4,
            bg="#2C3E50", fg="white", font=("Arial", 9, "bold"),
            relief="flat", bd=0, padx=4, pady=6
        ).grid(row=0, column=0, sticky="nsew", padx=1, pady=1)

        for c_idx, nome in enumerate(colunas, start=1):
            tk.Label(self.frame_tabela, text=nome,
                width=self._largura(nome)//7,
                bg=self._cor_cab(nome), fg="white", font=("Arial", 9, "bold"),
                relief="flat", bd=0, padx=6, pady=6,
                wraplength=self._largura(nome)-8
            ).grid(row=0, column=c_idx, sticky="nsew", padx=1, pady=1)

        # Linhas
        for r_idx, (df_row_idx, row) in enumerate(self.df.iterrows(), start=1):
            tk.Label(self.frame_tabela, text=str(r_idx), width=4,
                bg="#ECF0F1", fg="#7F8C8D", font=("Arial", 8),
                relief="flat", bd=0, padx=4
            ).grid(row=r_idx, column=0, sticky="nsew", padx=1, pady=1)

            for c_idx, nome_col in enumerate(colunas, start=1):
                val = row[nome_col]
                texto = str(val) if str(val) not in ("nan", "None") else ""
                bg = self._cor_cel(nome_col, r_idx)

                if nome_col == COLUNA_TAXA_ESCOLHA:
                    # Dropdown (Combobox) direto na célula
                    var = tk.StringVar(value=texto if texto in OPCOES_TAXA else "Sem taxa")
                    cb = ttk.Combobox(self.frame_tabela, textvariable=var,
                        values=OPCOES_TAXA, state="readonly",
                        width=self._largura(nome_col)//8, font=("Arial", 9)
                    )
                    cb.grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)

                    # Quando muda, atualiza df e recalcula Valor Taxa
                    def ao_mudar(event, row_i=df_row_idx, v=var, c=c_idx):
                        escolha = v.get()
                        self.df.at[row_i, COLUNA_TAXA_ESCOLHA] = escolha
                        novo_val = calcular_valor_taxa(self.df, row_i)
                        self.df.at[row_i, COLUNA_TAXA_VALOR] = novo_val
                        # Atualiza o Label do Valor Taxa
                        col_vt = list(self.df.columns).index(COLUNA_TAXA_VALOR)
                        lbl_vt = self._cells.get((row_i, col_vt))
                        if lbl_vt:
                            lbl_vt.config(text=f"{novo_val:.2f}" if novo_val else "0.00")

                    cb.bind("<<ComboboxSelected>>", ao_mudar)
                    self._cells[(df_row_idx, c_idx-1)] = cb

                else:
                    lbl = tk.Label(self.frame_tabela, text=texto,
                        width=self._largura(nome_col)//7,
                        bg=bg, fg=UI["texto"], font=("Arial", 9),
                        relief="flat", bd=0, padx=6, pady=4,
                        anchor="w" if nome_col == "Nome do produto" else "center"
                    )
                    lbl.grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)

                    if nome_col != COLUNA_TAXA_VALOR:
                        lbl.bind("<Double-Button-1>",
                            lambda e, r=df_row_idx, c=c_idx-1, w=lbl: self._iniciar_edicao(r, c, w))

                    self._cells[(df_row_idx, c_idx-1)] = lbl

    # ── Edição ──────────────────────────────────

    def _iniciar_edicao(self, row_idx, col_idx, lbl_widget):
        if self._celula_editando:
            self._confirmar_edicao()

        val_atual = str(self.df.iloc[row_idx, col_idx])
        if val_atual in ("nan", "None"): val_atual = ""

        entry = tk.Entry(self.frame_tabela, font=("Arial", 9),
            bg="#FFFDE7", fg="#1A2A3A", relief="solid", bd=1,
            highlightthickness=2, highlightcolor="#1565C0", highlightbackground="#1565C0"
        )
        entry.insert(0, val_atual)
        entry.select_range(0, tk.END)

        info = lbl_widget.grid_info()
        entry.grid(row=info["row"], column=info["column"], sticky="nsew", padx=1, pady=1)
        entry.focus_set()

        self._celula_editando = (row_idx, col_idx, entry, lbl_widget)
        entry.bind("<Return>",   lambda e: self._confirmar_edicao())
        entry.bind("<Escape>",   lambda e: self._cancelar_edicao())
        entry.bind("<Tab>",      lambda e: self._confirmar_edicao())
        entry.bind("<FocusOut>", lambda e: self._confirmar_edicao())

    def _confirmar_edicao(self):
        if not self._celula_editando:
            return
        row_idx, col_idx, entry, lbl_widget = self._celula_editando
        self._celula_editando = None
        novo_val = entry.get()
        entry.destroy()
        self.df.iloc[row_idx, col_idx] = novo_val
        lbl_widget.config(text=novo_val)
        lbl_widget.grid()

    def _cancelar_edicao(self):
        if not self._celula_editando:
            return
        _, _, entry, lbl_widget = self._celula_editando
        self._celula_editando = None
        entry.destroy()
        lbl_widget.grid()

    # ── Salvar ──────────────────────────────────

    def _salvar(self):
        if self._celula_editando:
            self._confirmar_edicao()
        try:
            salvar_xlsx(self.df, self.caminho_saida)
            self.cb_salvo(len(self.df))
            messagebox.showinfo("Salvo!", f"Planilha salva com sucesso!\n\n{self.caminho_saida}", parent=self)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar:\n{e}", parent=self)


# ─────────────────────────────────────────────
# JANELA PRINCIPAL
# ─────────────────────────────────────────────

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Extrator ML – Mercado Livre")
        self.root.geometry("680x590")
        self.root.resizable(False, False)
        self.root.configure(bg=UI["bg_claro"])
        self.arquivo_entrada = tk.StringVar()
        self.arquivo_saida   = tk.StringVar()
        self.df_atual        = None
        self._construir_interface()

    def _card(self, parent, pady=(6, 0)):
        outer = tk.Frame(parent, bg=UI["borda"])
        outer.pack(fill="x", padx=28, pady=pady)
        inner = tk.Frame(outer, bg=UI["bg_card"], padx=16, pady=12)
        inner.pack(fill="x", padx=1, pady=1)
        return inner

    def _num_label(self, parent, numero, titulo):
        row = tk.Frame(parent, bg=UI["bg_card"])
        row.pack(fill="x", pady=(0, 8))
        tk.Label(row, text=f" {numero} ", font=("Arial", 9, "bold"),
            bg=UI["azul"], fg="white"
        ).pack(side="left", padx=(0, 8))
        tk.Label(row, text=titulo, font=("Arial", 10, "bold"),
            bg=UI["bg_card"], fg=UI["texto"]
        ).pack(side="left")

    def _entry_row(self, parent, var, btn_texto, cmd):
        row = tk.Frame(parent, bg=UI["bg_card"])
        row.pack(fill="x")
        e = tk.Entry(row, textvariable=var, font=("Arial", 9),
            state="readonly", bg="#F0F4F8", fg=UI["texto"],
            relief="flat", highlightthickness=1,
            highlightbackground=UI["borda"], readonlybackground="#F0F4F8"
        )
        e.pack(side="left", fill="x", expand=True, ipady=6, padx=(0, 8))
        b = tk.Button(row, text=btn_texto, command=cmd,
            bg=UI["azul"], fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=12, pady=5
        )
        b.pack(side="left")
        _hover(b, UI["azul"], UI["azul_hover"])

    def _construir_interface(self):
        # Header
        hdr = tk.Frame(self.root, bg=UI["bg_escuro"], height=82)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📊  Extrator de Dados",
            font=("Arial", 17, "bold"), bg=UI["bg_escuro"], fg="white"
        ).place(relx=0.5, rely=0.36, anchor="center")
        tk.Label(hdr, text="Mercado Livre  /  Mercado Shops",
            font=("Arial", 9), bg=UI["bg_escuro"], fg="#90CAF9"
        ).place(relx=0.5, rely=0.70, anchor="center")

        tk.Frame(self.root, bg=UI["bg_claro"], height=14).pack(fill="x")

        c1 = self._card(self.root)
        self._num_label(c1, "1", "Planilha de origem  (Mercado Livre)")
        self._entry_row(c1, self.arquivo_entrada, "  Procurar  ", self._selecionar_entrada)

        c2 = self._card(self.root)
        self._num_label(c2, "2", "Onde salvar a planilha nova")
        self._entry_row(c2, self.arquivo_saida, "  Salvar em  ", self._selecionar_saida)

        # Tags das colunas
        c3 = self._card(self.root, pady=(6, 6))
        tk.Label(c3, text="Colunas que serão geradas na planilha:",
            font=("Arial", 9, "bold"), bg=UI["bg_card"], fg=UI["subtexto"]
        ).pack(anchor="w", pady=(0, 7))

        tags = tk.Frame(c3, bg=UI["bg_card"])
        tags.pack(fill="x")

        tags_info = [
            ("Nome do produto",    UI["tag_bg"],   UI["tag_fg"]),
            ("Receita",            UI["tag_bg"],   UI["tag_fg"]),
            ("Tarifa 12% ou 17%",  UI["tag_bg"],   UI["tag_fg"]),
            ("Receita por envio",  UI["tag_bg"],   UI["tag_fg"]),
            ("Tarifa de envio",    UI["tag_bg"],   UI["tag_fg"]),
            ("Desconto bônus",     UI["tag_bg"],   UI["tag_fg"]),
            ("Total sobre",        UI["tag_bg"],   UI["tag_fg"]),
            (COLUNA_ESTADO,        "#E8DAEF",      "#6C3483"),
            (COLUNA_ICMS,          "#FDEBD0",      "#B7410E"),
            (COLUNA_TAXA_ESCOLHA,  "#D6EAF8",      "#1A5276"),
            (COLUNA_TAXA_VALOR,    "#D1F2EB",      "#154360"),
            (COLUNA_GUBEE,         UI["tag_gubee"],UI["tag_gubee_fg"]),
        ]

        # Duas linhas de tags
        linha1 = tk.Frame(c3, bg=UI["bg_card"])
        linha1.pack(anchor="w", pady=1)
        linha2 = tk.Frame(c3, bg=UI["bg_card"])
        linha2.pack(anchor="w", pady=1)

        for i, (nome, bg, fg) in enumerate(tags_info):
            frame_destino = linha1 if i < 7 else linha2
            tk.Label(frame_destino, text=nome, font=("Arial", 8, "bold"),
                bg=bg, fg=fg, padx=7, pady=3
            ).pack(side="left", padx=2, pady=1)

        # Botões
        btn_frame = tk.Frame(self.root, bg=UI["bg_claro"])
        btn_frame.pack(pady=16)

        self.btn_extrair = tk.Button(btn_frame, text="▶   Extrair Dados",
            command=self._processar,
            bg=UI["verde"], fg="white", font=("Arial", 13, "bold"),
            relief="flat", padx=28, pady=10, cursor="hand2"
        )
        self.btn_extrair.pack(side="left", padx=8)
        _hover(self.btn_extrair, UI["verde"], UI["verde_hover"])

        self.btn_tabela = tk.Button(btn_frame, text="📋   Ver / Editar Tabela",
            command=self._abrir_tabela,
            bg=UI["azul"], fg="white", font=("Arial", 11, "bold"),
            relief="flat", padx=18, pady=10, cursor="hand2",
            state="disabled"
        )
        self.btn_tabela.pack(side="left", padx=8)
        _hover(self.btn_tabela, UI["azul"], UI["azul_hover"])

        # Status bar
        status_bar = tk.Frame(self.root, bg="#E8EDF2", height=32)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        self.label_status = tk.Label(status_bar,
            text="  ●  Aguardando arquivo...",
            font=("Arial", 9), bg="#E8EDF2", fg=UI["subtexto"], anchor="w"
        )
        self.label_status.pack(side="left", fill="y", padx=8)

    def _selecionar_entrada(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha do Mercado Livre",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        if caminho:
            self.arquivo_entrada.set(caminho)
            pasta     = os.path.dirname(caminho)
            nome_base = os.path.splitext(os.path.basename(caminho))[0]
            self.arquivo_saida.set(os.path.join(pasta, f"{nome_base}_EXTRAIDO.xlsx"))
            self.label_status.config(text="  ●  Arquivo selecionado.", fg="#2E7D32")

    def _selecionar_saida(self):
        caminho = filedialog.asksaveasfilename(
            title="Salvar planilha nova como...",
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")]
        )
        if caminho:
            self.arquivo_saida.set(caminho)

    def _processar(self):
        entrada = self.arquivo_entrada.get()
        saida   = self.arquivo_saida.get()
        if not entrada:
            messagebox.showwarning("Atenção", "Selecione a planilha de origem primeiro.")
            return
        if not saida:
            messagebox.showwarning("Atenção", "Escolha onde salvar a planilha nova.")
            return

        self.btn_extrair.config(state="disabled", text="⏳  Processando...")
        self.root.update()

        try:
            self.label_status.config(text="  ●  Lendo arquivo...", fg=UI["subtexto"])
            self.root.update()
            df, faltando = preparar_dados(entrada)
            self.df_atual = df

            self.label_status.config(text="  ●  Salvando planilha...", fg=UI["subtexto"])
            self.root.update()
            salvar_xlsx(df, saida)

            self.label_status.config(text=f"  ●  Concluído! {len(df)} linhas extraídas.", fg="#2E7D32")
            self.btn_tabela.config(state="normal")

            aviso = ""
            if faltando:
                aviso = "\n\n⚠️ Colunas não encontradas:\n" + "\n".join(f"  • {c}" for c in faltando)
            messagebox.showinfo("Concluído!",
                f"Planilha criada com sucesso!\n\n{len(df)} vendas extraídas.\n\nSalvo em:\n{saida}{aviso}")

        except Exception as e:
            self.label_status.config(text="  ●  Erro ao processar.", fg="#C0392B")
            messagebox.showerror("Erro", f"Ocorreu um erro:\n\n{str(e)}")
        finally:
            self.btn_extrair.config(state="normal", text="▶   Extrair Dados")

    def _abrir_tabela(self):
        if self.df_atual is None:
            return
        def ao_salvar(qtd):
            self.label_status.config(text=f"  ●  Tabela salva ({qtd} linhas).", fg="#2E7D32")
        JanelaTabela(self.root, self.df_atual, self.arquivo_saida.get(), ao_salvar)


# ─────────────────────────────────────────────
if __name__ == "__main__":
    janela = tk.Tk()
    App(janela)
    janela.mainloop()
